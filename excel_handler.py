import logging
import pandas as pd
from io import BytesIO
from typing import Dict, List, Tuple, Optional
import openpyxl

logger = logging.getLogger(__name__)


def get_existing_content(file, sheet_name: str, content_column: str) -> Dict[int, str]:
    """
    Read existing content from a target sheet.
    Returns a dict mapping row index (0-based) to existing content.
    Only includes rows that have non-empty content.
    """
    existing = {}
    try:
        wb = openpyxl.load_workbook(file)
        file.seek(0)

        if sheet_name not in wb.sheetnames:
            return existing

        ws = wb[sheet_name]

        # Find the content column index
        content_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == content_column:
                content_col_idx = col_idx
                break

        if content_col_idx is None:
            content_col_idx = 2  # Default to column B

        # Read existing content (starting from row 2, skipping header)
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=content_col_idx).value
            if cell_value is not None and str(cell_value).strip() != "":
                existing[row_idx - 2] = str(cell_value)  # 0-based index

    except Exception as e:
        logger.warning(f"Error reading existing content from '{sheet_name}': {e}")

    return existing


def read_source_with_fallback(
    file,
    primary_sheet: str,
    fallback_sheet: str,
    content_column: str,
) -> Tuple[List[str], List[int]]:
    """
    Read content from primary sheet, fill empty cells from fallback sheet.

    Returns:
        (content_list, fallback_used_indices) — the merged content and which indices used fallback.
    """
    # Read primary sheet
    primary_df = pd.read_excel(file, sheet_name=primary_sheet, engine="openpyxl")
    file.seek(0)

    if content_column not in primary_df.columns:
        logger.warning(f"Column '{content_column}' not found in '{primary_sheet}'")
        return list(primary_df.iloc[:, 1] if primary_df.shape[1] > 1 else primary_df.iloc[:, 0]), []

    primary_content = primary_df[content_column].tolist()

    # Try to read fallback sheet
    fallback_used = []
    try:
        fallback_df = pd.read_excel(file, sheet_name=fallback_sheet, engine="openpyxl")
        file.seek(0)

        if content_column in fallback_df.columns:
            fallback_content = fallback_df[content_column].tolist()

            for i in range(len(primary_content)):
                val = primary_content[i]
                if val is None or str(val).strip() == "" or str(val).strip().lower() == "nan":
                    if i < len(fallback_content):
                        fb_val = fallback_content[i]
                        if fb_val is not None and str(fb_val).strip() != "" and str(fb_val).strip().lower() != "nan":
                            primary_content[i] = fb_val
                            fallback_used.append(i)
                            logger.info(f"Row {i}: used fallback from '{fallback_sheet}'")
    except Exception as e:
        logger.warning(f"Could not read fallback sheet '{fallback_sheet}': {e}")
        file.seek(0)

    return primary_content, fallback_used


def get_workbook_info(file) -> Tuple[List[str], Dict[str, List[str]], Dict[str, int]]:
    """
    Get information about an Excel workbook.
    Returns (sheet_names, columns_per_sheet, rows_per_sheet).
    """
    wb = openpyxl.load_workbook(file)
    file.seek(0)

    sheet_names = wb.sheetnames
    columns_per_sheet = {}
    rows_per_sheet = {}

    for sheet_name in sheet_names:
        df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
        file.seek(0)
        columns_per_sheet[sheet_name] = list(df.columns)
        rows_per_sheet[sheet_name] = len(df)

    return sheet_names, columns_per_sheet, rows_per_sheet


def read_source_sheet(file, sheet_name: str) -> pd.DataFrame:
    """Read the source sheet containing content to translate."""
    df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
    file.seek(0)
    return df


def create_translated_workbook(
    original_file,
    source_sheet: str,
    content_column: str,
    translations: Dict[str, List[str]]
) -> BytesIO:
    """
    Create a new workbook with translations filled in, preserving original structure.
    """
    original_file.seek(0)
    wb = openpyxl.load_workbook(original_file)
    original_file.seek(0)

    source_ws = wb[source_sheet]

    # Find the content column index (1-based in openpyxl)
    content_col_idx = None
    for col_idx, cell in enumerate(source_ws[1], start=1):
        if cell.value == content_column:
            content_col_idx = col_idx
            break

    if content_col_idx is None:
        content_col_idx = 2

    # Fill in translations for each target sheet
    for sheet_name, translated_content in translations.items():
        if sheet_name in wb.sheetnames and sheet_name != source_sheet:
            ws = wb[sheet_name]

            ws.cell(row=1, column=content_col_idx).value = content_column

            for row_idx, content in enumerate(translated_content, start=2):
                ws.cell(row=row_idx, column=content_col_idx).value = content

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_multi_file_zip(files: Dict[str, BytesIO]) -> BytesIO:
    """Create a ZIP file containing multiple Excel files."""
    import zipfile

    output = BytesIO()

    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, file_buffer in files.items():
            file_buffer.seek(0)
            zf.writestr(filename, file_buffer.read())

    output.seek(0)
    return output
